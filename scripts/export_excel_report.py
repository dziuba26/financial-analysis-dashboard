"""Export Apple's financial analysis into a polished Excel workbook."""

from pathlib import Path
import sqlite3
import sys

import pandas as pd
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.append(str(PROJECT_ROOT))

from scripts.calculate_metrics import PROCESSED_DATA_PATH
from scripts.load_financial_data import load_apple_financials
from scripts.load_to_sqlite import DATABASE_PATH

REPORT_PATH = PROJECT_ROOT / "reports" / "apple_financial_analysis_2022_2025.xlsx"
SQL_QUERY_PATHS = {
    "Yearly Results": PROJECT_ROOT / "sql" / "yearly_results.sql",
    "Highest Revenue Year": PROJECT_ROOT / "sql" / "highest_revenue_year.sql",
    "Average Revenue Growth": PROJECT_ROOT / "sql" / "average_revenue_growth.sql",
    "Highest ROE": PROJECT_ROOT / "sql" / "highest_roe.sql",
    "Highest EPS": PROJECT_ROOT / "sql" / "highest_eps.sql",
}


def load_processed_metrics() -> pd.DataFrame:
    """Load the processed metrics CSV for the Excel report."""
    metrics = pd.read_csv(PROCESSED_DATA_PATH)
    metrics["period_end"] = pd.to_datetime(metrics["period_end"]).dt.date
    return metrics


def run_sql_query(query_path: Path) -> pd.DataFrame:
    """Run a saved SQL query and return its results."""
    query = query_path.read_text()
    with sqlite3.connect(DATABASE_PATH) as connection:
        return pd.read_sql_query(query, connection)


def build_dashboard_data(metrics: pd.DataFrame) -> pd.DataFrame:
    """Create a compact dashboard summary from the latest fiscal year."""
    latest_year = metrics.sort_values("fiscal_year").iloc[-1]

    return pd.DataFrame(
        {
            "Metric": [
                "Fiscal Year",
                "Revenue",
                "Net Income",
                "Diluted EPS",
                "ROE",
                "Revenue Growth",
                "Net Profit Margin",
                "Free Cash Flow",
            ],
            "Value": [
                int(latest_year["fiscal_year"]),
                latest_year["revenue_usd_millions"],
                latest_year["net_income_usd_millions"],
                latest_year["diluted_eps_usd_per_share"],
                latest_year["return_on_equity"],
                latest_year["revenue_growth_rate"],
                latest_year["net_profit_margin"],
                latest_year["free_cash_flow_usd_millions"],
            ],
            "Unit": [
                "",
                "USD millions",
                "USD millions",
                "USD/share",
                "%",
                "%",
                "%",
                "USD millions",
            ],
        }
    )


def autosize_columns(worksheet) -> None:
    """Adjust Excel column widths to fit the worksheet content."""
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))

        worksheet.column_dimensions[column_letter].width = min(max_length + 2, 45)


def style_worksheet(worksheet) -> None:
    """Apply simple finance-report formatting to an Excel worksheet."""
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    worksheet.freeze_panes = "A2"
    autosize_columns(worksheet)


def format_metric_columns(workbook) -> None:
    """Format ratios and dollar values for easier review in Excel."""
    dashboard = workbook["Dashboard"]
    for row in range(2, dashboard.max_row + 1):
        unit = dashboard.cell(row=row, column=3).value
        value_cell = dashboard.cell(row=row, column=2)
        if unit == "%":
            value_cell.number_format = "0.00%"
        elif unit == "USD/share":
            value_cell.number_format = "$0.00"
        elif unit == "USD millions":
            value_cell.number_format = "$#,##0"

    processed = workbook["Processed Metrics"]
    percent_columns = {
        "revenue_growth_rate",
        "net_income_growth_rate",
        "gross_margin",
        "operating_margin",
        "net_profit_margin",
        "free_cash_flow_margin",
        "return_on_equity",
    }
    for column in range(1, processed.max_column + 1):
        header = processed.cell(row=1, column=column).value
        if header in percent_columns:
            for row in range(2, processed.max_row + 1):
                processed.cell(row=row, column=column).number_format = "0.00%"


def add_dashboard_charts(workbook) -> None:
    """Add Excel line charts that mirror the core dashboard trends."""
    dashboard = workbook["Dashboard"]
    metrics = workbook["Processed Metrics"]

    chart_specs = [
        ("Revenue Trend", 3, "E2"),
        ("Net Income Trend", 6, "E18"),
        ("Diluted EPS Trend", 7, "M2"),
        ("ROE Trend", 29, "M18"),
    ]

    years = Reference(metrics, min_col=1, min_row=2, max_row=metrics.max_row)
    for title, value_column, anchor in chart_specs:
        chart = LineChart()
        chart.title = title
        chart.y_axis.title = title
        chart.x_axis.title = "Fiscal Year"
        data = Reference(metrics, min_col=value_column, min_row=1, max_row=metrics.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(years)
        chart.height = 7
        chart.width = 12
        dashboard.add_chart(chart, anchor)


def write_sql_summary_sheet(writer: pd.ExcelWriter) -> None:
    """Write each saved SQL query result into one summary worksheet."""
    start_row = 0
    sheet_name = "SQL Summary"

    for title, query_path in SQL_QUERY_PATHS.items():
        result = run_sql_query(query_path)
        result.to_excel(writer, sheet_name=sheet_name, startrow=start_row + 1, index=False)

        worksheet = writer.sheets[sheet_name]
        title_cell = worksheet.cell(row=start_row + 1, column=1)
        title_cell.value = title
        title_cell.font = Font(bold=True, size=12)

        start_row += len(result) + 4


def export_excel_report(report_path: Path = REPORT_PATH) -> Path:
    """Create the Excel workbook used as the project's Excel deliverable."""
    raw_data = load_apple_financials()
    raw_data["period_end"] = raw_data["period_end"].dt.date
    processed_metrics = load_processed_metrics()
    dashboard_data = build_dashboard_data(processed_metrics)

    report_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(report_path, engine="openpyxl") as writer:
        dashboard_data.to_excel(writer, sheet_name="Dashboard", index=False)
        raw_data.to_excel(writer, sheet_name="Raw Data", index=False)
        processed_metrics.to_excel(writer, sheet_name="Processed Metrics", index=False)
        write_sql_summary_sheet(writer)

        for worksheet in writer.book.worksheets:
            style_worksheet(worksheet)

        format_metric_columns(writer.book)
        add_dashboard_charts(writer.book)

    return report_path


if __name__ == "__main__":
    saved_report_path = export_excel_report()
    print(f"Saved Excel report to: {saved_report_path}")
